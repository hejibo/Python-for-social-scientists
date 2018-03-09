# -*- coding: utf-8 -*-
"""
Created on Sun May 17 22:28:23 2015

@author: Jibo

Crawl number of review pages


http://www.amazon.com/gp/cdp/member-reviews/A2D1LPEUCTNT8X?ie=UTF8&display=public&page=1&sort_by=MostRecentReview


sample reviewer #:
    A2D1LPEUCTNT8X
    A1E1LEVQ9VQNK

    http://www.amazon.com/gp/cdp/member-reviews/A1E1LEVQ9VQNK?ie=UTF8&display=public&page=1&sort_by=MostRecentReview

"""


import multiprocessing as mp
import random
import string
import threading
from time import ctime,sleep

import urllib
from BeautifulSoup import BeautifulSoup
import urllib2
import cPickle as p

def CrawlPage(PageIndex):
    try:
        url ='http://www.amazon.com/gp/cdp/member-reviews/A3TWBW1B17R151?ie=UTF8&display=public&page=%s&sort_by=MostRecentReview'%PageIndex
        pageFile = urllib.urlopen(url)
        pageHtml = pageFile.read()
        pageFile.close()
        reviewfile =r'D:\OneDrive\Research\amazon online review\Step 3. crawl number of review pages\Reviewer First Page Cache\amazon-product-review-%d-page.data'%PageIndex
        # the name of the file where we will store the object
        # Write to the file

        f = file(reviewfile, 'w')
        p1.dump(pageHtml, f) # dump the object to a file
        f.close()
        print 'finished page:'%PageIndex
    except:
        print '!!!!!!!!!!!!!!!!!!!!!!!!failed for %d page'%PageIndex
#
#for PageIndex in xrange(1,1000):
#    #PageIndex=1
#    CrawlPage(PageIndex)

def CrawlPageTest(PageIndex,reviewerID):

    url ='http://www.amazon.com/gp/cdp/member-reviews/%s?ie=UTF8&display=public&page=%s&sort_by=MostRecentReview'%("A2KBF2OYR359AJ",9)
    print url
    pageFile = urllib.urlopen(url)
    pageHtml = pageFile.read()
    pageFile.close()
    reviewfile = r"G:\Step 4. crawl  all review pages\All review page cache\A2KBF2OYR359AJ\reviewer-A2KBF2OYR359AJ-9-page.data"
    # the name of the file where we will store the object
    # Write to the file

    f = file(reviewfile, 'w')
    p1.dump(pageHtml, f) # dump the object to a file
    f.close()

def ReadReviewerIDFile():
    filename = "amazon-reviewer-list-review-count-part1.txt"
    infile = open(filename,'r')
    reviewIDList =[ line.split() for line in infile.readlines()]
    infile.close()
    return reviewIDList

def WriteDataCSV(outputArray,filename):
    outfile = open(filename,'w')
    rowIndex=0

    print >>outfile,"rankIndexValue","\t", 	"name" ,	"\t","reviewerID" ,"\t",	"TotalReviews" ,"\t",	"HelpfulVotes" ,"\t",	"crNumPercentHelpful" ,"\t",	"crNumFanVoters","\t","productName","\t","price","\t","productLink","\t","starReview","\t","reviewTime","\t","isVerifiedPurchase","\t","IsVineReviewFreeProduct","\t","reviewText"

    #print >>outfile,'rankIndexValue','\t','name','\t','reviewerID','\t','TotalReviews','\t','HelpfulVotes','\t','crNumPercentHelpful','\t','crNumFanVoters'
    for (rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters,productName,price,productLink,starReview,reviewTime,isVerifiedPurchase,IsVineReviewFreeProduct,reviewText) in outputArray:
        try:
            print >>outfile,rankIndexValue, "\t",	name ,"\t",	reviewerID ,"\t",	TotalReviews ,"\t",	HelpfulVotes ,	"\t",crNumPercentHelpful ,"\t",	crNumFanVoters,"\t",productName,"\t",price,"\t",productLink,"\t",starReview,"\t",reviewTime,"\t",isVerifiedPurchase,"\t",IsVineReviewFreeProduct,"\t",reviewText.encode('utf8')
        except:
            print 'cannot print'

    outfile.close()

def LoadCachedPage(reviewerID,PageIndex):

    reviewfile = r"/Users/hejibo/Downloads/Step 4. crawl  all review pages/All review page cache/%s/reviewer-%s-%s-page.data"%(reviewerID,reviewerID,PageIndex)
    # the name of the file where we will store the object
    # Write to the file

    f = file(reviewfile)
    soup = p.load(f)
    f.close()
    return soup

def LoadCachedPageFromFile(reviewfile):

    #reviewfile = r"/Users/hejibo/Downloads/Step 4. crawl  all review pages/All review page cache/%s/reviewer-%s-%s-page.data"%(reviewerID,reviewerID,PageIndex)
    # the name of the file where we will store the object
    # Write to the file

    f = file(reviewfile)
    soup = p.load(f)
    f.close()
    return soup

def getInfo(rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters,PageIndex):
    '''
    get text between links
    http://stackoverflow.com/questions/6251319/extract-text-between-link-tags-in-python-using-beautifulsoup
    '''

    #sample
    # the following will get "<span class="h3color tiny">This review is from: </span>,
    #print soup
    #soupParsed = BeautifulSoup(soup)
    #print soupParsed
    soup =  LoadCachedPage(reviewerID,PageIndex)
    productInfo = parseSourceCode(soup,rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters)
    return productInfo

def parseSourceCode(soup,rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters):
    soupParsed = BeautifulSoup(soup)

    # the second table is for the review content
    datapart= soupParsed.findAll("table",{"class":"small"})
    # review part is marked by "<div style="margin-left:0.5em;">"
    reviewParts= soupParsed.findAll("div",{"style":"margin-left:0.5em;"})

    # for the review section of each product, there are two table with class value of small
    # so table 1, 4, 7, 10 is for the product link
    # in the product section, if the product is a vine free product, there is no price tag.
    productInfo=[]
    for productIndex in range(0,10):
        try:
            #productIndex=1
            productPart = datapart[productIndex*3+1]
            productName= productPart.text
            price = productName[productName.find("Price:")+6:]
            price.replace("$","")
            productName = productName[:productName.find("Price:")]

            #print len(datapart)
            #print "----------------------------------------------"
            #print productPart
            productLink = productPart.find("a")['href']
            #print productLink
            #productInfo.append([productName,productLink])

            #reviewPartsIndex=0
            reviewPart= reviewParts[productIndex]
            starReview=reviewPart.find("img")["title"]

            #print starReview
            reviewTime = reviewPart.find("nobr").text

            isVerifiedPurchase="Not Verified"

            CountVerifiedPurchase = reviewPart.findAll("span",{"class":"crVerifiedStripe"})
            if CountVerifiedPurchase:
                #isVerifiedPurchase = reviewPart.find("span",{"class":"crVerifiedStripe"}).text
                isVerifiedPurchase ="Verified Purchase"

            print 'isVerifiedPurchase:',isVerifiedPurchase,'\n'

            #print reviewTime
            AllReviewText =  reviewParts[productIndex].text
            if "Vine Customer Review of Free Product" in AllReviewText:
                IsVineReviewFreeProduct="YesVineReviewFreeProduct"
            else:
                IsVineReviewFreeProduct="NoVineReviewFreeProduct"
            #print IsVineReviewFreeProduct
            reviewText=reviewPart.find("div",{"class":"reviewText"}).text
            print "productName,",productName,"\n"
            print "productLink:\n", productLink,"\n"
            print "starReview:\n",starReview,"\n"
            print "reviewTime:\n",reviewTime,"\n"
            print "IsVineReviewFreeProduct:\n",IsVineReviewFreeProduct,"\n"
            print reviewText
            print "------------------------------------------------------------"
            productInfo.append([rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters,productName,price,productLink,starReview,reviewTime,isVerifiedPurchase,IsVineReviewFreeProduct,reviewText])
        except:
            print 'failed this product section'
        #productReviewInfo=zip(productInfo,reviewInfo)
        #break
    return productInfo

def ParseMultiplePageThreads(reviewerID,pageStart,pageEnd):
    threads = []

    # Setup a list of processes that we want to run
    #reviewerID = 'A1E1LEVQ9VQNK'
    #processes = [mp.Process(target=CrawlPageTest, args=(x, reviewerID)) for x in range(4)]

    for pageIndex in range(pageStart,pageEnd):
        t1 = threading.Thread(target=CrawlPageTest,args=(pageIndex,reviewerID,))
        threads.append(t1)

    for t in threads:
        t.setDaemon(True)
        t.start()

def ParseMultiplePageProcesses(rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters,pageStart,pageEnd):

    from multiprocessing.pool import ThreadPool
    pool = ThreadPool(processes=30)


    return_val=[]

    for PageIndex in range(pageStart,pageEnd):
        async_result = pool.apply_async(getInfo, (rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters, PageIndex))
        return_val.extend( async_result.get())
    print return_val
    return return_val

def maxPageCounts(reviewCount):
    '''
    calculate the total number of pages
    '''
    reviewCount = int(reviewCount)
    if reviewCount%10==0:
        pageCounts = reviewCount/10
    else:
        pageCounts = int(reviewCount/10)+1
    return pageCounts


def ReadReviewerProfile():
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'amazon-reviewer-list-complete-profile.xlsx')
    sheet_ranges = wb['Sheet1']
    nrows = len(sheet_ranges.rows)

    profileList =[]
    for rowIndex in range(1,nrows):
        #rowIndex =1
        try:
            rankIndexValue=sheet_ranges.cell(row = rowIndex, column = 0).value
            name=sheet_ranges.cell(row = rowIndex, column = 1).value
            reviewerID=sheet_ranges.cell(row = rowIndex, column = 2).value.strip()
            #print reviewerID
            TotalReviews=int(sheet_ranges.cell(row = rowIndex, column = 3).value)
            #print TotalReviews
            HelpfulVotes=sheet_ranges.cell(row = rowIndex, column = 4).value
            crNumPercentHelpful=sheet_ranges.cell(row = rowIndex, column = 5).value
            crNumFanVoters=sheet_ranges.cell(row = rowIndex, column = 6).value

            #print rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters
            profileList.append([rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters])
        except:
            continue

    #print(sheet_ranges['D18'].value)

    return profileList

def ensure_dir(path):
    '''ensure a directy is valid for creating a file, if the directory does not exist,
    create the directory'''
    import os
    d = os.path.dirname(path)
    if not os.path.exists(d):
        os.makedirs(d)

def ParseAllReviewInfo():

    #print reviewIDList
    productReviewInfo=[]


    #reviewIDList= ReadReviewerIDFile()
    #print reviewIDList

    profileList =ReadReviewerProfile()
    print "profileList:\n",profileList

    counts =0
    for (rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters) in profileList:
        #reviewerID = 'A1E1LEVQ9VQNK'
        PageIndex = 0

        print reviewerID,TotalReviews
        #reviewerID = 'A2KBF2OYR359AJ'
        #PageIndex = 9
        print reviewerID
        counts =counts +1
        print '-----------------------------------------analyze-----------------------'
        print "reviewerID:",reviewerID
        print '---------------------------------------------------------------------'

        maxPages =  maxPageCounts(TotalReviews)
        crawlStep=30
        for indexPage in range(0,maxPages,crawlStep):
            print indexPage

            #soup= LoadCachedPage(reviewerID,PageIndex)
            #productInfo = getInfo(reviewerID,PageIndex)

            #productInfo=ParseMultiplePageProcesses(reviewerID,indexPage,min(indexPage+crawlStep,maxPages))
            #productReviewInfo.extend(productInfo)
            try:
                productInfo=ParseMultiplePageProcesses(rankIndexValue, 	name ,	reviewerID ,	TotalReviews ,	HelpfulVotes ,	crNumPercentHelpful ,	crNumFanVoters,indexPage,min(indexPage+crawlStep,maxPages))
                productReviewInfo.extend(productInfo)
            #print productInfo
            except:
                continue

            #break
        filename = r'/Users/hejibo/Downloads/Step 4. crawl  all review pages/output folder/amazon-review-'+reviewerID+".csv"
        rootpath =r'/Users/hejibo/Downloads/Step 4. crawl  all review pages/output folder/'
        ensure_dir(rootpath)
        WriteDataCSV(productReviewInfo,filename)
        productReviewInfo=[]
        #WriteDataCSV(productReviewInfo,filename)
    print "-_-!!"

def ParseSamplePage():
    #reviewfile =r'/Users/hejibo/Downloads/Step 4. crawl  all review pages/sample pages/reviewer-ACJT8MUC0LRF0-0-page.data'

    reviewfile =r'/Users/hejibo/Downloads/Step 4. crawl  all review pages/sample pages/reviewer-A1UQNX8S78JU0G-0-page.data'

    soup = LoadCachedPageFromFile(reviewfile)
    productInfo = parseSourceCode(soup,"1", 	"test name" ,	"A1A0EYOH42GUH1" ,	"13" ,	"25,115" ,	"96%" ,	"34")
    print productInfo

    reviewerID="ACJT8MUC0LRF0"
    filename = r'/Users/hejibo/Downloads/Step 4. crawl  all review pages/output folder/amazon-review-'+reviewerID+".csv"
    rootpath =r'/Users/hejibo/Downloads/Step 4. crawl  all review pages/output folder/'
    ensure_dir(rootpath)
    WriteDataCSV(productInfo,filename)
    #getInfo(reviewerID,PageIndex)

#ParseAllReviewInfo()
ParseSamplePage()

